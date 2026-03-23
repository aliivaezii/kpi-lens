"""
Excel report exporter — produces a styled multi-sheet workbook for weekly KPI review.

Design decisions:
- Three sheets: Executive Summary, Anomaly Detail, KPI Trends.
  Analysts open the Summary sheet first; drill-down lives in the other two.
- openpyxl PatternFill is used for RAG status cells because conditional
  formatting in openpyxl is write-only and not readable by downstream tools.
- Column widths are auto-sized by measuring cell content rather than using
  openpyxl's unreliable auto-size, which requires LibreOffice to render.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from kpi_lens.kpis.definitions import (
    KPI_BY_NAME,  # noqa: F401 — reserved for future per-KPI formatting
)

# RAG status → hex fill colour (no leading #; openpyxl uses ARGB "FF..." prefix)
_STATUS_FILL: dict[str, PatternFill] = {
    "green": PatternFill("solid", fgColor="FF92D050"),
    "yellow": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF2F4F8F")


def _auto_width(ws: Any) -> None:
    """Set column widths based on the longest cell content in each column."""
    for col_cells in ws.columns:
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        # 1.2 multiplier adds padding; cap at 50 to avoid oversized columns
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_length * 1.2 + 2, 50
        )


def _write_header_row(ws: Any, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def build_summary_sheet(ws: Any, snapshot: dict[str, Any]) -> None:
    """Populate the Executive Summary sheet from an enriched snapshot dict."""
    ws.title = "Executive Summary"
    _write_header_row(
        ws,
        [
            "KPI",
            "Display Name",
            "Latest Value",
            "Unit",
            "Status",
            "vs Benchmark (%)",
            "Direction",
        ],
    )
    for kpi_name, data in snapshot.items():
        status = data.get("health_status", "")
        row = [
            kpi_name,
            data.get("display_name", kpi_name),
            data.get("value"),
            data.get("unit", ""),
            status.upper() if status else "",
            data.get("benchmark_distance"),
            data.get("direction", ""),
        ]
        ws.append(row)
        status_cell = ws.cell(row=ws.max_row, column=5)
        if status in _STATUS_FILL:
            status_cell.fill = _STATUS_FILL[status]
    _auto_width(ws)


def build_anomaly_sheet(ws: Any, anomalies: list[dict[str, Any]]) -> None:
    """Populate the Anomaly Detail sheet from a list of anomaly dicts."""
    ws.title = "Anomaly Detail"
    _write_header_row(
        ws,
        [
            "ID",
            "KPI",
            "Period Start",
            "Period End",
            "Observed",
            "Expected Low",
            "Expected High",
            "Severity",
            "Detector",
            "Entity",
            "Acknowledged",
        ],
    )
    for anomaly in anomalies:
        severity = anomaly.get("severity", 0)
        ws.append(
            [
                anomaly.get("id"),
                anomaly.get("kpi_name"),
                str(anomaly.get("period_start", "")),
                str(anomaly.get("period_end", "")),
                anomaly.get("observed_value"),
                anomaly.get("expected_low"),
                anomaly.get("expected_high"),
                round(severity, 3),
                anomaly.get("detector_name"),
                anomaly.get("entity"),
                "Yes" if anomaly.get("is_acknowledged") else "No",
            ]
        )
        # Colour the entire row by severity
        sev_fill = (
            _STATUS_FILL["red"]
            if severity >= 0.7
            else _STATUS_FILL["yellow"]
            if severity >= 0.4
            else _STATUS_FILL["green"]
        )
        for cell in ws[ws.max_row]:
            cell.fill = sev_fill
    _auto_width(ws)


def build_trends_sheet(
    ws: Any,
    trends: dict[str, list[dict[str, Any]]],
) -> None:
    """
    Populate the KPI Trends sheet.

    `trends` maps kpi_name → list of {"period_end": ..., "value": ...} dicts,
    the same shape returned by repository.get_kpi_series().
    """
    ws.title = "KPI Trends"
    # Pivot: one column per KPI, one row per week
    # Collect all period_end dates across all KPIs for the index column
    all_dates: set[str] = set()
    for series in trends.values():
        for row in series:
            all_dates.add(str(row.get("period_end", "")))
    sorted_dates = sorted(all_dates)

    kpi_names = list(trends.keys())
    _write_header_row(ws, ["Period End"] + kpi_names)

    date_value_map: dict[str, dict[str, Any]] = {d: {} for d in sorted_dates}
    for kpi_name, series in trends.items():
        for row in series:
            period_key = str(row.get("period_end", ""))
            date_value_map[period_key][kpi_name] = row.get("value")

    for period_date in sorted_dates:
        ws.append(
            [period_date] + [date_value_map[period_date].get(k) for k in kpi_names]
        )
    _auto_width(ws)


def generate_workbook(
    snapshot: dict[str, Any],
    anomalies: list[dict[str, Any]],
    trends: dict[str, list[dict[str, Any]]],
) -> bytes:
    """
    Generate a complete KPI-Lens Excel workbook and return its bytes.

    All three arguments come from the repository layer — never call the
    repository from within this function (keep reporting decoupled from DB).
    """
    wb = Workbook()
    build_summary_sheet(wb.active, snapshot)
    build_anomaly_sheet(wb.create_sheet(), anomalies)
    build_trends_sheet(wb.create_sheet(), trends)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
